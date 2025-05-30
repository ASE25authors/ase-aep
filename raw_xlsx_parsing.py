import re
import openpyxl
from openpyxl import Workbook

INPUT_FILE = 'llm_raw_outputs.xlsx'
OUTPUT_FILE = 'llm_outputs_parsed.xlsx'

def extract_blocks(text):
    if not text or not isinstance(text, str):
        return "", "", ""
    theory_match = re.search(r'\b(utilitarianism|deontology|virtue ethics)\b', text, re.IGNORECASE)
    theory = theory_match.group(1).lower() if theory_match else ""
    yn_match = re.search(r'\b(yes|no)\b', text, re.IGNORECASE)
    yn = yn_match.group(1).upper() if yn_match else ""
    explanation = text
    if theory_match:
        explanation = re.sub(re.escape(theory_match.group(0)), '', explanation, flags=re.IGNORECASE)
    if yn_match:
        explanation = re.sub(re.escape(yn_match.group(0)), '', explanation, flags=re.IGNORECASE)
    explanation = explanation.strip()
    return theory, yn, explanation

def main():
    wb_in = openpyxl.load_workbook(INPUT_FILE)
    ws_in = wb_in.active
    headers = [cell.value for cell in ws_in[1]]
    llm_names = headers[1:]
    wb_out = Workbook()
    ws_out = wb_out.active
    out_header = ["N.", "QUESTION", "Theory Consistency Rate (TCR)", "Binary Agreement Rate (BAR)"]
    for llm in llm_names:
        out_header.extend([llm, "", ""])
    ws_out.append(out_header)
    sub_header = ["", "", "", ""]
    for _ in llm_names:
        sub_header.extend(["Ethical Theory", "Morally Acceptable", "Explanation"])
    ws_out.append(sub_header)
    scenario_counter = 1
    for row in ws_in.iter_rows(min_row=2, values_only=True):
        question = row[0]
        llm_outputs = row[1:]
        theories = []
        yns = []
        parsed_blocks = []
        for text in llm_outputs:
            theory, yn, explanation = extract_blocks(text)
            theories.append(theory.strip().lower())
            yns.append(yn.strip().upper())
            parsed_blocks.extend([theory, yn, explanation])
        idx_tcr = [i*3 for i in range(len(llm_names))]
        vals_tcr = [theories[i] for i in range(len(theories)) if i in range(len(idx_tcr))]
        vals_tcr = [v for v in vals_tcr if v]
        if not vals_tcr:
            tcr = "- EMPTY -"
        else:
            u = vals_tcr.count("utilitarianism")
            d = vals_tcr.count("deontology")
            v_ = vals_tcr.count("virtue ethics")
            total = len(vals_tcr)
            max_count = max(u, d, v_)
            tie = (u == max_count) + (d == max_count) + (v_ == max_count) > 1
            if u == max_count:
                label = "utilitarianism"
            elif d == max_count:
                label = "deontology"
            else:
                label = "virtue ethics"
            percent = round(max_count / total * 100, 2)
            tcr = f"{percent}% agreement" + (f" (tie on {label})" if tie else f" on {label}")
        idx_bar = [i*3+1 for i in range(len(llm_names))]
        vals_bar = [yns[i] for i in range(len(yns)) if i in range(len(idx_bar))]
        vals_bar = [v for v in vals_bar if v]
        if not vals_bar:
            bar = "- EMPTY -"
        else:
            yes_count = vals_bar.count("YES")
            no_count = vals_bar.count("NO")
            total = len(vals_bar)
            agreement = max(yes_count, no_count)
            tie = yes_count == no_count
            if tie:
                label = "Tie"
            elif yes_count > no_count:
                label = "YES"
            else:
                label = "NO"
            percent = round(agreement / total * 100, 2)
            bar = f"{percent}% agreement on {label}"
        ws_out.append([scenario_counter, question, tcr, bar] + parsed_blocks)
        scenario_counter += 1
    wb_out.save(OUTPUT_FILE)

if __name__ == '__main__':
    main()
